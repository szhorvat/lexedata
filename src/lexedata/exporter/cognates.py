# -*- coding: utf-8 -*-
import re
import abc
import typing as t
import urllib.parse
from pathlib import Path

import pycldf
import openpyxl as op

from lexedata import types, cli, util
from lexedata.util import parse_segment_slices


WARNING = "\u26A0"

# ----------- Remark: Indices in excel are always 1-based. -----------

# TODO: Make comments on Languages, Cognatesets, and Judgements appear as notes
# in Excel.

# Type aliases, for clarity
CognatesetID = str


class BaseExcelWriter:
    """Class logic for matrix-shaped Excel export."""

    row_table: str
    header: t.List[t.Tuple[str, str]]

    def __init__(
        self,
        dataset: pycldf.Dataset,
        database_url: t.Optional[str] = None,
        logger: cli.logging.Logger = cli.logger,
    ):
        self.dataset = dataset
        self.set_header()

        self.row_id = self.dataset[self.row_table, "id"].name

        self.URL_BASE = database_url

        self.wb = op.Workbook()
        self.ws: op.worksheet.worksheet.Worksheet = self.wb.active

        self.logger = logger

    def create_excel(
        self,
        out: Path,
        size_sort: bool = False,
        rows: t.Optional[types.RowObject] = None,
        language_order="name",
    ) -> None:
        """Convert the initial CLDF into an Excel cognate view

        The Excel file has columns "CogSet", one column each mirroring the
        other cognateset metadata, and then one column per language.

        The rows contain cognate data. If a language has multiple reflexes in
        the same cognateset, these appear in different cells, one below the
        other.

        Parameters
        ==========
        out: The path of the Excel file to be written.

        size_sort: If true, cognatesets are ordered by the number of cognates
            corresponding to the cognateset

        language_order: column name, languages appear ordered by given column name from
            LanguageTable

        """
        # cldf names

        c_name = self.dataset["LanguageTable", "name"].name
        c_id = self.dataset["LanguageTable", "id"].name
        c_form_id = self.dataset["FormTable", "id"].name
        c_form_concept_reference = self.dataset["FormTable", "parameterReference"].name

        # Define the columns, i.e. languages and write to excel
        self.lan_dict: t.Dict[str, int] = {}
        excel_header = [name for cldf, name in self.header]
        # TODO: wrap the following two blocks into a
        # get_sorted_languages() -> t.OrderedDict[languageReference, Column Header/Titel/Name]
        # function
        if language_order:
            c_sort = self.dataset["LanguageTable", f"{language_order}"].name
            languages = sorted(
                self.dataset["LanguageTable"], key=lambda x: x[c_sort], reverse=False
            )
        else:
            # sorted returns a list, so better return a list here as well
            languages = list(self.dataset["LanguageTable"])
        for col, lan in cli.tq(
            enumerate(languages, len(excel_header) + 1),
            task="Writing languages to excel header",
            total=len(languages),
        ):
            # TODO: This should be based on the foreign key relation
            self.lan_dict[lan[c_id]] = col
            excel_header.append(lan[c_name])
        self.ws.append(excel_header)

        # map form_id to id of associated concept
        concept_id_by_form_id = dict()
        for f in self.dataset["FormTable"]:
            concept = f[c_form_concept_reference]
            if isinstance(concept, str):
                concept_id_by_form_id[f[c_form_id]] = concept
            else:
                concept_id_by_form_id[f[c_form_id]] = concept[0]

        # Again, row_index 2 is indeed row 2, row 1 is header
        row_index = 1 + 1

        if rows is None:
            rows = self.collect_rows()

        all_judgements = self.collect_forms_by_row()

        # iterate over all rows
        for cogset in cli.tq(
            rows,
            task="Writing rows to Excel",
            total=len(rows),
        ):
            # possibly a cogset can appear without any judgment, if so ignore it
            if cogset["id"] not in all_judgements:
                continue
            # write all forms of this cognateset to excel
            new_row_index = self.create_formcells(
                cogset,
                all_judgements[cogset["id"]],
                row_index,
            )
            # write rows for cognatesets
            for row in range(row_index, new_row_index):
                self.write_row_header(cogset, row)

            row_index = new_row_index

        self.after_filling(row_index)
        self.wb.save(filename=out)

    def create_formcells(
        self,
        cogset: types.CogSet,
        all_forms: t.Dict[str, types.Form],
        row_index: int,
    ) -> int:
        """Writes all forms for given cognate set to Excel.

        Take all forms for a given cognate set as given by the database, create
        a hyperlink cell for each form, and write those into rows starting at
        row_index.

        Return the row number of the first empty row after this cognate set,
        which can then be filled by the following cognate set.

        """
        # Read the forms from the database and group them by language
        forms = t.DefaultDict[int, t.List[types.Form]](list)
        for form in all_forms:
            forms[self.lan_dict[form["languageReference"]]].append(form)

        if not forms:
            return row_index + 1

        # maximum of rows to be added
        maximum_cogset = max([len(c) for c in forms.values()])
        for column, cells in forms.items():
            for row, judgement in enumerate(cells, row_index):
                self.create_formcell(judgement, column, row)
        # increase row_index and return
        row_index += maximum_cogset

        return row_index

    def create_formcell(self, judgement, column: int, row: int) -> None:
        """Fill the given cell with the form's data.

        In the cell described by ws, column, row, dump the data for the form:
        Write into the the form data, and supply a comment from the judgement
        if there is one.

        """
        cell_value = self.form_to_cell_value(judgement)
        form_cell = self.ws.cell(row=row, column=column, value=cell_value)
        comment = judgement.get("comment")
        if comment:
            form_cell.comment = op.comments.Comment(comment, __package__)
        if self.URL_BASE:
            link = self.URL_BASE.format(urllib.parse.quote(judgement["id"]))
            form_cell.hyperlink = link

    def collect_rows(self):
        return util.cache_table(self.dataset, self.row_table).values()

    @abc.abstractmethod
    def after_filling(self, row_index):
        "What should happen after the last regular row has been written?"


class ExcelWriter(BaseExcelWriter):
    """Class logic for cognateset Excel export."""

    row_table = "CognatesetTable"

    def __init__(
        self,
        dataset: pycldf.Dataset,
        database_url: t.Optional[str] = None,
        singleton_cognate: bool = False,
        singleton_status: t.Optional[str] = None,
        logger: cli.logging.Logger = cli.logger,
    ):
        super().__init__(dataset=dataset, database_url=database_url, logger=logger)
        # assert that all required tables are present in Dataset
        try:
            for _ in dataset["CognatesetTable"]:
                break
        except (KeyError, FileNotFoundError):
            cli.Exit.INVALID_DATASET(
                "This script presupposes a separate CognatesetTable. Call `lexedata.edit.add_table CognatesetTable` to automatically add one."
            )
        try:
            for _ in dataset["CognateTable"]:
                break
        except (KeyError, FileNotFoundError):
            cli.Exit.NO_COGNATETABLE(
                "This script presupposes a separate CognateTable. Call `lexedata.edit.add_cognate_table` to automatically add one."
            )
        self.singleton = singleton_cognate
        self.singleton_status = singleton_status
        self.row_id = self.dataset["CognatesetTable", "id"].name
        if self.singleton_status is not None:
            if ("Status_Column", "Status_Column") not in self.header:
                self.logger.warning(
                    f"You requested that I set the status of new singleton cognate sets to {self.singleton_status}, but your CognatesetTable has no Status_Column to write it to. If you want a Status "
                )

    def write_row_header(self, cogset, row_number: int):
        try:
            c_comment = self.dataset["CognatesetTable", "comment"].name
        except KeyError:
            c_comment = None
        for col, (db_name, header) in enumerate(self.header, 1):
            # db_name is '' when add_central_concepts is activated
            # and there is no concept column in cognateset table
            # else read value from cognateset table
            if header == "Central_Concept" and db_name == "":
                # this is the concept associated to the first cognate in this cognateset
                raise NotImplementedError(
                    "You expect central conceps in your cognate set table, but you don't have any central concepts stored with your cognate sets"
                )
            else:
                if db_name == "":
                    continue
                column = self.dataset[self.row_table, db_name]
                if column.separator is None:
                    value = cogset[db_name]
                else:
                    value = column.separator.join([str(v) for v in cogset[db_name]])
            cell = self.ws.cell(row=row_number, column=col, value=value)
            # Transfer the cognateset comment to the first Excel cell.
            if c_comment and col == 1 and cogset.get(c_comment):
                cell.comment = op.comments.Comment(
                    re.sub(f"-?{__package__}", "", cogset[c_comment] or "").strip(),
                    "lexedata.exporter",
                )

    def after_filling(self, row_index):
        if not self.singleton:
            return
        # write remaining forms to singleton congatesets if switch is activated

        all_forms = util.cache_table(self.dataset)

        # remove all forms that appear in judgements
        for k in cli.tq(
            self.collect_forms_by_row().values(),
            task="Writing singleton cognatesets to excel",
            total=row_index,
        ):
            for form in k:
                all_forms.pop(form["id"], None)
        # create for remaining forms singleton cognatesets and write to file
        for i, form_id in enumerate(all_forms):
            # write form to file
            form = all_forms[form_id]
            self.create_formcell(
                form,
                self.lan_dict[form["languageReference"]],
                row_index,
            )
            # write singleton cognateset to excel
            for col, (db_name, header) in enumerate(self.header, 1):
                if db_name == "id":
                    value = f"X{i+1}_{form['languageReference']}"
                elif db_name == "name":
                    value = form_id
                elif db_name == "parameterReference":
                    value = all_forms[form_id]["parameterReference"]
                elif header == "Status_Column" and self.singleton_status is not None:
                    value = self.singleton_status
                else:
                    value = ""
                self.ws.cell(row=row_index, column=col, value=value)
            row_index += 1

    def set_header(self):
        c_id = self.dataset["CognatesetTable", "id"].name
        try:
            c_comment = self.dataset["CognatesetTable", "comment"].name
        except (KeyError):
            c_comment = None
        self.header = []
        for column in self.dataset["CognatesetTable"].tableSchema.columns:
            if column.name == c_id:
                self.header.insert(0, ("id", "CogSet"))
            elif column.name == c_comment:
                continue
            else:
                try:
                    property = util.cldf_property(column.propertyUrl)
                except AttributeError:
                    property = column.name
                self.header.append((property, column.name))

    def form_to_cell_value(self, form: types.Form) -> str:
        """Build a string describing the form itself

        Provide the best transcription and all translations of the form strung
        together.

        k a w e n a t j a k a
        d i +dúpe
        +iíté+ k h ú
        tákeː+toː
        h o n _tiem_litimuleni
        hont i e m _litimuleni
        """
        segments = self.get_segments(form)
        if not segments:
            transcription = form["form"]
        else:
            transcription = ""
            # TODO: use CLDF property instead of column name
            included_segments: t.Iterable[int]
            try:
                included_segments = set(
                    parse_segment_slices(form["segmentSlice"], enforce_ordered=True)
                )
            except KeyError:
                included_segments = range(len(form["segments"]))
            except ValueError:
                # What if segments overlap or cross? Overlap shouldn't happen,
                # but we don't check here. Crossing might happen, but this
                # serialization cannot reflect it, so we enforce order,
                # expecting that an error message here will be more useful than
                # silently messing with data. If the check fails, we take the
                # whole segment and warn.
                self.logger.warning(
                    f"In form {form['id']}, with judgement{form['judgement_id']}, segment slice {form['segmentSlice']} is invalid."
                )
                included_segments = range(len(form["segments"]))

            included = False
            for i, s in enumerate(segments):
                if included and i not in included_segments:
                    transcription += " }" + s
                    included = False
                elif not included and i in included_segments:
                    transcription += "{ " + s
                    included = True
                elif i in included_segments:
                    transcription += " " + s
                else:
                    transcription += s
            if included:
                transcription += " }"

            transcription = transcription.strip()
        translations = []

        suffix = ""
        try:
            c_comment = self.dataset["FormTable", "comment"].name
            if form.get(c_comment):
                suffix = f" {WARNING:}"
        except (KeyError):
            pass

        # corresponding concepts
        # (multiple concepts) and others (single concept)
        if isinstance(form["parameterReference"], list):
            for f in form["parameterReference"]:
                translations.append(f)
        else:
            translations.append(form["parameterReference"])
        return "{:} ‘{:}’{:}".format(transcription, ", ".join(translations), suffix)

    def get_segments(self, form: types.Form):
        try:
            return form["segments"]
        except (KeyError):
            self.logger.warning("No segments column found. Falling back to cldf form.")
            return form["form"]

    def collect_forms_by_row(
        self,
    ) -> t.Mapping[types.Cognateset_ID, t.List[types.Form]]:
        forms: t.Mapping[types.Form_ID, types.Form] = util.cache_table(
            self.dataset
        )  # TODO: index_column = reference target of CognateTable's formReference
        all_forms: t.MutableMapping[
            types.Cognateset_ID, t.List[types.Form]
        ] = t.DefaultDict(list)
        c_j_cognateset = self.dataset["CognateTable", "cognatesetReference"].name
        try:
            c_j_slice = self.dataset["CognateTable", "segmentSlice"].name
        except KeyError:
            c_j_slice = None
        try:
            c_j_comment = self.dataset["CognateTable", "comment"].name
        except KeyError:
            c_j_comment = None
        c_j_form = self.dataset["CognateTable", "formReference"].name
        for judgement in self.dataset["CognateTable"]:
            form_with_judgement_metadata: types.Form = types.Form(
                forms[judgement[c_j_form]]
            )
            form_with_judgement_metadata["cognatesetReference"] = judgement[
                c_j_cognateset
            ]
            form_with_judgement_metadata["segmentSlice"] = judgement.get(c_j_slice)
            form_with_judgement_metadata["formComment"] = form_with_judgement_metadata[
                "comment"
            ]
            form_with_judgement_metadata["comment"] = judgement.get(c_j_comment)
            # TODO: Add other printable judgement properties
            all_forms[judgement[c_j_cognateset]].append(form_with_judgement_metadata)
        return all_forms


if __name__ == "__main__":
    parser = cli.parser(description="Create an Excel cognate view from a CLDF dataset")
    parser.add_argument(
        "excel",
        type=Path,
        help="File path for the generated cognate excel file.",
    )
    parser.add_argument(
        "--size-sort",
        action="store_true",
        default=False,
        help="List the biggest cognatesets first (within a group, if another sort order is specified by --sort-cognatesets-by)",
    )
    parser.add_argument(
        "--sort-languages-by",
        help="The name of a column in the LanguageTable to sort languages by in the output",
    )
    parser.add_argument(
        "--sort-cognatesets-by",
        help="The name of a column in the CognatesetTable to sort cognates by in the output",
        default="id",
    )
    parser.add_argument(
        "--url-template",
        type=str,
        default="https://example.org/lexicon/{:}",
        help="A template string for URLs pointing to individual forms. For example, to"
        " point to lexibank, you would use https://lexibank.clld.org/values/{:}."
        " (default: https://example.org/lexicon/{:})",
    )
    parser.add_argument(
        "--add-singletons-with-status",
        default=None,
        metavar="MESSAGE",
        help="Include in the output all forms that don't belong to a cognateset. For each form, a singleton cognateset is created, and its status column (if there is one) is set to MESSAGE.",
    )
    parser.add_argument(
        "--add-singletons",
        action="store_const",
        const="automatic singleton",
        help="Short for `--add-singletons-with-status='automatic singleton'`",
        dest="add_singletons_with_status",
    )
    args = parser.parse_args()
    logger = cli.setup_logging(args)

    dataset = pycldf.Wordlist.from_metadata(args.metadata)
    try:
        cogsets = list(dataset["CognatesetTable"])
    except (KeyError):
        cli.Exit.INVALID_DATASET(
            "Dataset has no explicit CognatesetTable. Add one using `lexedata.edit.add_table CognatesetTable`."
        )

    E = ExcelWriter(
        dataset,
        database_url=args.url_template,
        singleton_cognate=args.add_singletons_with_status is None,
        singleton_status=args.add_singletons_with_status,
        logger=logger,
    )

    try:
        cogset_order = E.dataset["CognatesetTable", args.sort_cognatesets_by].name
    except (KeyError):
        cli.Exit.INVALID_COLUMN_NAME(
            f"No column '{args.sort_cognatesets_by}' in your CognatesetTable."
        )

    cogsets = list(dataset["CognatesetTable"])
    # Sort first by size, then by the specified column, so that if both
    # happen, the cognatesets are globally sorted by the specified column
    # and within one group by size.
    if args.size_sort:
        raise NotImplementedError
        cogsets.sort(
            key=lambda x: ...,
            reverse=True,
        )

    cogsets.sort(key=lambda c: c[cogset_order])

    E.create_excel(
        args.excel,
        size_sort=args.size_sort,
        rows=cogsets,
        language_order=args.sort_languages_by,
    )
